import datetime as dt

from openpyxl import Workbook
from django.db.models import Count

from .models import Robot


class RobotsReport:

    def __init__(self, start_date=None, end_date=None):
        if end_date:
            self.end_date = end_date
        else:
            self.end_date = dt.datetime.now()

        if start_date:
            self.start_date = start_date
        else:
            self.start_date = self.end_date - dt.timedelta(weeks=1)

    def get_report(self):
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        unique_models = Robot.objects.filter(
            created__range=(self.start_date, self.end_date)
        ).values_list(
            'model', flat=True
        ).distinct()

        for model in unique_models:
            sheet = wb.create_sheet(title=model)
            sheet.cell(row=1, column=1, value="Модель")
            sheet.cell(row=1, column=2, value="Версия")
            sheet.cell(row=1, column=3, value="Количество за неделю")

            robots = Robot.objects.filter(
                model=model,
                created__range=(self.start_date, self.end_date)
            ).values(
                'version'
            ).annotate(
                count=Count('version')
            ).order_by(
                'version')

            for ids, record in enumerate(robots):
                sheet.cell(row=2+ids, column=1, value=model)
                sheet.cell(row=2+ids, column=2, value=record['version'])
                sheet.cell(row=2+ids, column=3, value=record['count'])

        excel_file_path = "robots_report.xlsx"  # Замените на путь к файлу, куда хотите сохранить Excel-таблицу
        wb.save(excel_file_path)
        return excel_file_path


